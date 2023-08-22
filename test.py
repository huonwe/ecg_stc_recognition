import os

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import argparse
from utils_config import get_config, get_model, prepare_test, prepare_check

# import torchvision.ops.focal_loss as FocalLoss
from sklearn.metrics import precision_recall_curve, average_precision_score, f1_score
import matplotlib.pyplot as plt
import matplotlib
from loss import *

import openpyxl
# from torchviz import make_dot
# from tensorboardX import SummaryWriter

songTi = matplotlib.font_manager.FontProperties(fname='/home/huonwe/hpsimplifiedhans-regular.ttf')
check_time = 10


def main(args):
    device = torch.device("cuda:" + args.ctx if torch.cuda.is_available() else "cpu")
    cfg = get_config(args.network)
    print(device)
    # print(cfg)
    test_loader = prepare_test(
        batch_size=50
    )
    # model_path = (
    #     "./models/model7500/200/model_3.pth"
    # )
    model_path = (
        "./models/model7500/200/model_3.pth"
    )
    
    score = test(model_path,test_loader,device,cfg)
    print(score)
    return
    check_loader = prepare_check(batch_size=50)
    check(model_path,check_loader,device,cfg)
    

@torch.no_grad()
def test(model_path, test_loader, device, cfg, marker="test"):
    network = get_model(cfg, num_channel=12)
    network.load_state_dict(torch.load(model_path))
    network = network.eval()
    predict_mat = []
    predict_mat_sigmoid = []
    label_mat = []
    plt.figure(figsize=(5,5),dpi=200)
    plt.ion()
    for _, v_data in enumerate(test_loader):
        inputs, labels = v_data
        # inputs, labels = inputs.to(device), labels.to(device)
        outputs = network(inputs)
        # outputs = F.sigmoid(outputs)
        outputs_sigmoid = F.sigmoid(outputs)
        predict_mat.extend(outputs)
        predict_mat_sigmoid.extend(outputs_sigmoid)
        label_mat.extend(labels)
    #     for idx,ecg in enumerate(inputs):
    #         plt.cla()
    #         for c in ecg:
    #             plt.plot(np.arange(0,len(c)),c)
    #         this_output = outputs[idx]
    #         this_output_sigmoid = outputs_sigmoid[idx]
    #         this_label = labels[idx]
    #         # score = f"score: STE:{this_output[0]} STD:{this_output[1]} Others:{this_output[2]}\n"
    #         # posi = f"posibility: STE:{this_output_sigmoid[0]} STD:{this_output_sigmoid[1]} Others:{this_output_sigmoid[2]}\n"
    #         # lab = f"label: {this_label[0]} {this_label[1]} {this_label[2]}"
    #         score = "score: {:>8.2f} {:>8.2f} {:>8.2f}\n".format(this_output[0],this_output[1],this_output[2])
    #         posi = "posibility: {:>8.2%} {:>8.2%} {:>8.2%}\n".format(this_output_sigmoid[0],this_output_sigmoid[1],this_output_sigmoid[2])
    #         truth = "truth: {:>10.0f} {:>10.0f} {:>10.0f}".format(this_label[0],this_label[1],this_label[2])
    #         plt.title(score+posi+truth)
    #         plt.pause(0.6)
        
    # plt.ioff()
    # plt.show()
        
    hist_mat_001 = []
    hist_mat_001_v = []
    hist_mat_110 = []
    hist_mat_110_v = []
    hist_mat_100 = []
    hist_mat_100_v = []
    hist_mat_010 = []
    hist_mat_010_v = []
    
    predict_mat = predict_mat_sigmoid
    for idx_label, label in enumerate(label_mat):
        if label[2] == 1:
            hist_mat_001.append(predict_mat[idx_label][2].item())
        if label[2] == 0:
            hist_mat_001_v.append(predict_mat[idx_label][2].item())
        if label[0] == 1:
            hist_mat_100.append(predict_mat[idx_label][0].item())
        else:
            hist_mat_100_v.append(predict_mat[idx_label][0].item())
        if label[1] == 1:
            hist_mat_010.append(predict_mat[idx_label][1].item())
        else:
            hist_mat_010_v.append(predict_mat[idx_label][1].item())
    print(np.mean(hist_mat_100),np.mean(hist_mat_100_v))
    print(np.std(hist_mat_100),np.std(hist_mat_100_v))
    # plt.figure(dpi=150)
    # plt.subplots_adjust(wspace =0, hspace =0.5)
    # plt.subplot(2,1,1),plt.title("测试集STD标签为真的模型预测值分布",fontproperties=songTi,fontsize=10),plt.hist(hist_mat_010,bins=20,rwidth=1,stacked=False)
    # plt.ylabel("frequency")
    # plt.xlim(-15,15)
    # plt.subplot(2,1,2),plt.title("测试集STD标签为假的模型预测值分布",fontproperties=songTi,fontsize=10),plt.hist(hist_mat_010_v,bins=20,rwidth=1,stacked=False)
    # plt.xlabel("predict value (before sigmoid)")
    # plt.ylabel("frequency")
    # plt.xlim(-15,15)
    # plt.show()
    
    # plt.figure(dpi=150)
    # plt.subplots_adjust(wspace =0, hspace =0.5)
    # plt.subplot(2,1,1),plt.title("测试集Others标签为真的模型预测值分布",fontproperties=songTi,fontsize=10),plt.hist(hist_mat_001,bins=20,rwidth=1,stacked=False)
    # plt.ylabel("frequency")
    # plt.xlim(0,1)
    # plt.subplot(2,1,2),plt.title("测试集Others标签为真的模型预测值分布",fontproperties=songTi,fontsize=10),plt.hist(hist_mat_001_v,bins=20,rwidth=1,stacked=False)
    # plt.xlabel("predict value (after sigmoid)")
    # plt.ylabel("frequency")
    # plt.xlim(0,1)
    # plt.show()
    
    
    
    
    
    
    
    # return
    count11 = 0
    count22 = 0
    count33 = 0
    count100 = 0
    count010 = 0
    count110 = 0
    for l in label_mat:
        if l[0] == 1:
            count11 += 1
        if l[1] == 1:
            count22 += 1
        if l[2] == 1:
            count33 += 1
        if l[1] == 1 and l[0] == 1:
            count110 += 1
    print(count11,count22,count33, count110)
    # return
    score1 = auprc(predict_mat, label_mat, 0, cfg,"STE")
    score2 = auprc(predict_mat, label_mat, 1, cfg,"STD")
    score3 = auprc(predict_mat, label_mat, 2, cfg,"Others")

    
    result = {}
    result['score'] = [score1,score2,score3]
    return result


def pred2Lable(pred, threshold=0.5) -> list:
    # pred = F.sigmoid(pred)
    pred_truth = [0, 0, 0]
    if pred[0] > threshold:
        pred_truth[0] = 1
    if pred[1] > threshold:
        pred_truth[1] = 1
    if pred[2] > threshold:
        pred_truth[2] = 1
    if pred_truth[0] or pred_truth[1]:
        pred_truth[2] = 0
    if pred_truth[2] == 1:
        pred_truth[0] = 0
        pred_truth[1] = 0
    return pred_truth


def pacc(ouputs, labels) -> float:
    correct = 0
    length = len(labels)
    for idx, pred in enumerate(ouputs):
        pred_truth = pred2Lable(pred)
        if pred_truth == [x.item() for x in labels[idx]]:
            correct += 1

    return correct / length


def auprc(predict_mat, label_mat, dim, cfg, marker=""):
    truth = np.array([x[dim].item() for x in label_mat])
    try:
        pred = np.array([x[dim].item() for x in predict_mat])
    except Exception:
        pred = np.array([x[dim] for x in predict_mat])

    precision, recall, threholds = precision_recall_curve(truth, pred, pos_label=1)
    print(marker)
    for i in range(0,len(precision)):
        if precision[i] > 0.6 and recall[i] > 0.6:
            print(f"threhold:{threholds[i]},precision:{precision[i]},recall:{recall[i]}")
    score = average_precision_score(truth, pred, pos_label=1)
    plt.xlim(0, 1.05)
    plt.ylim(0, 1.05)
    plt.xlabel("Recall")
    plt.ylabel("Precision")
    plt.scatter(recall, precision, c="r")
    plt.title("PRC-"+marker)
    plt.savefig(f"./auprc/{cfg.network}-{marker}-{dim}.jpg")
    plt.cla()
    return score

# import netron
@torch.no_grad()
def check(model_path, check_loader, device, cfg):
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet('sheet1',0)
    
    network = get_model(cfg, num_channel=12).to(device)
    network.load_state_dict(torch.load(model_path))
    network = network.eval()
    predict_mat = []
    label_mat = []
    # sample_date = torch.randn((1,12,250)).cuda()
    # output = network(sample_date)
    
    # torch.onnx.export(network,sample_date,"./log/model.pth")
    # # netron.start("./log/model.pth")
    # with SummaryWriter("./log", comment="sample_model_visualization") as sw:
    #     sw.add_graph(network, sample_date)
    # torch.save(network, "./log/modelviz.pt")
    # g = make_dot(output)
    # g.render('modelviz',view=False)
    # return
    for _, v_data in enumerate(check_loader):
        inputs, names = v_data
        inputs = inputs.to(device)
        outputs = network(inputs)

        predict_mat.extend(F.sigmoid(outputs))
        label_mat.extend(names)
    
    idx = 0
    for pred, name in zip(predict_mat,label_mat):
        sheet.cell(row=idx+2,column=1).value=idx
        sheet.cell(row=idx+2,column=2).value=str(name.item())+".mat"
        sheet.cell(row=idx+2,column=3).value=pred[0].item()
        sheet.cell(row=idx+2,column=4).value=pred[1].item()
        sheet.cell(row=idx+2,column=5).value=pred[2].item()
        idx += 1
    excel.save("results/check.xlsx")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Train option")
    parser.add_argument("network", type=str, help="Pytorch ECG Traing")
    parser.add_argument("--ctx", default="0", type=str, help="cuda")
    parser.add_argument(
        "--set", default="7500-ori", type=str, help="hdf5 suffix"
    )

    main(parser.parse_args())
